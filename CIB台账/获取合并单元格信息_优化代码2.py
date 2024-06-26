import pandas as pd
from datetime import datetime
import numpy as np
import tkinter as tk
from tkinter import filedialog
import xlwings as xw
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from tqdm import tqdm

def select_excel_file():
    """选择Excel文件的对话框"""
    return filedialog.askopenfilename(title="选择Excel文件", filetypes=[("Excel Files", "*.xlsx;*.xls")])

def get_merge_info(cell):
    """获取合并单元格的信息"""
    if cell.api.MergeCells:
        merge_area = cell.api.MergeArea
        return (
            merge_area.Cells(1).Row,  # 起始行号
            merge_area.Cells(1).Column,  # 起始列号
            merge_area.Cells(merge_area.Rows.Count, 1).Row,  # 结束行号
            merge_area.Cells(1, merge_area.Columns.Count).Column  # 结束列号
        )
    else:
        return cell.row, cell.column, cell.row, cell.column

def extract_merge_info(sheet, start_row, start_col):
    """提取合并单元格的信息，并过滤掉Merge_Context为空的记录"""
    max_column = sheet.used_range.last_cell.column
    data = []
    for col_index in range(ord(start_col) - ord('A'), max_column + 1):
        col_letter = get_column_letter(col_index + 1)
        cell = sheet[f'{col_letter}{start_row}']
        try:
            start_row_num, start_col_num, end_row_num, end_col_num = get_merge_info(cell)
            start_address = f"{get_column_letter(start_col_num)}{start_row_num}"
            end_address = f"{get_column_letter(end_col_num)}{end_row_num}"
            merge_count = (end_col_num - start_col_num + 1) * (end_row_num - start_row_num + 1) if cell.api.MergeCells else 1
            merge_context = cell.value if cell.api.MergeCells else ''
            # 直接在这里过滤掉merge_context为空的记录
            if merge_context:
                data.append({
                    'Start_Address': start_address,
                    'End_Address': end_address,
                    'Start_Column': start_col_num,
                    'End_Column': end_col_num,
                    'Merged_Cell_Count': merge_count,
                    'Merge_Context': merge_context
                })
        except Exception as e:
            print(f"处理列 {col_letter} 时发生错误: {e}")
    return data

def get_data_from_range(sheet, header_range, start_col_num, end_col_num):
    start_col = xw.utils.col_name(start_col_num)
    end_col = xw.utils.col_name(end_col_num)
    headers = sheet.range(header_range).value
    start_row = 10
    last_row = sheet.cells.last_cell.row
    end_row = sheet.range(f'{end_col}{last_row}').end('up').row
    data_range = sheet.range(f'{start_col}{start_row}:{end_col}{end_row}')
    data = data_range.value
    df = pd.DataFrame(data, columns=headers[:len(data[0])])
    df.dropna(how='all', inplace=True)
    return df



def main():
    try:
        # 选择文件
        file_path = select_excel_file()
        if not file_path:
            print("没有选择文件。")
            return
        # 提取合并单元格信息
        data_TS = extract_merge_info(sht1, 6, 'Z')
        data_Car = extract_merge_info(sht1, 8, 'Z')
        # 将信息转换为DataFrame
        df_TS_info = pd.DataFrame(data_TS)
        df_Car_info = pd.DataFrame(data_Car)
        # 去除重复行并筛选非空的'Merge_Context'
        # df_info = df_info.drop_duplicates().dropna(subset=['Merge_Context'])
        # 打印结果
        print(df_TS_info)
        print(df_Car_info)

        with xw.Book(file_path) as wb:
            sht1 = wb.sheets['Sheet1']
            book_name = sht1.range('A1').value
            sht1.range('O6').value = "制造分部\nMF"

            current_date = datetime.now().strftime('%Y%m%d')

            df1_1 = get_data_from_range(sht1, 'A6:S6', 1, 19)
            df1_2 = get_data_from_range(sht1, 'W6:Y6', 23, 25)
            df1_3 = get_data_from_range(sht1, 'T8:U8', 20, 21)
            df1 = pd.concat([df1_1, df1_2, df1_3], axis=1)

            columns_to_drop = ['工艺\nMET', '外控\nSQC', '采购\nPROC', '运维\nO&M', '项目\nPM', '工程\nENG',
                            '售后质量', 'Engineer', '质量-外控', '售后质量', '采购', '运维', '项目', '工程']









        # 保存并关闭工作簿
        wb.save()
    except FileNotFoundError:
        print(f"文件找不到: {file_path}")
    except PermissionError:
        print(f"没有权限访问文件: {file_path}")
    except Exception as e:
        print(f"发生错误: {e}")
    finally:
        wb.close()
        xw.App().quit()

if __name__ == "__main__":
    main()