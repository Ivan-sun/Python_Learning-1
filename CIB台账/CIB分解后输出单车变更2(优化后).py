import pandas as pd
from datetime import datetime
import numpy as np
import tkinter as tk
from tkinter import filedialog
import xlwings as xw
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from tqdm import tqdm

def get_merge_info(cell):
    if cell.api.MergeCells:
        merge_area = cell.api.MergeArea
        return (
            merge_area.Cells(1).Row, 
            merge_area.Cells(1).Column, 
            merge_area.Cells(merge_area.Rows.Count, 1).Row, 
            merge_area.Cells(1, merge_area.Columns.Count).Column
        )
    else:
        return cell.row, cell.column, cell.row, cell.column

def apply_format(cell, font_name='等线', font_size=14, font_color='FF0000', is_bold=True, border=True, align_center=True):
    """应用格式到单元格"""
    font = Font(name=font_name, color=font_color, bold=is_bold, size=font_size)
    cell.font = font
    if border:
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        cell.border = thin_border
    if align_center:
        alignment = Alignment(horizontal='center', vertical='center')
        cell.alignment = alignment
        
# 添加遗漏的函数定义
def get_data_from_range(sheet, header_range, start_col_num, end_col_num):
    start_col = xw.utils.col_name(start_col_num)
    end_col = xw.utils.col_name(end_col_num)
    headers = sheet.range(header_range).value
    start_row = 10
    last_row = sheet.cells.last_cell.row
    end_row = sheet.range(f'{end_col}{last_row}').end('up').row
    data_range = sheet.range(f'{start_col}{start_row}:{end_col}{end_row}')
    data = data_range.value
    df = pd.DataFrame(data, columns=headers[:len(data[0])])
    df.dropna(how='all', inplace=True)
    return df

# 优化：使用函数封装重复的逻辑
def process_sheet(sheet, start_col_num, end_col_num, merge_context):
    df_i = get_data_from_range(sheet, 'Z9:AJ9', start_col_num, end_col_num)
    merged_df = pd.concat([df1.reset_index(drop=True), df_i.reset_index(drop=True)], axis=1)
    merged_df.drop(columns_to_drop, errors='ignore', inplace=True)
    sum_CN = merged_df.shape[0] - 1
    merged_df = merged_df[~(merged_df[['车间\nMF', '外协\nOT', '制造分部\nMF']] == '/').all(axis=1)]
    sum_MF_CN = merged_df.shape[0]
    merged_df = merged_df[merged_df['变更关闭状态'] == 'open']
    merged_df = merged_df[(merged_df[['生产', '外协\nOT', '制造分部\nMF']].apply(lambda x: x.str.contains('未完成')).any(axis=1))]
    sum_MF_OPEN_CN = merged_df.shape[0]
    return merged_df, sum_CN, sum_MF_CN, sum_MF_OPEN_CN

# 优化：使用with语句管理资源
try:
    file_path = filedialog.askopenfilename(title="选择Excel文件", filetypes=[("Excel Files", "*.xlsx;*.xls")])
    with xw.Book(file_path) as wb:
        sht1 = wb.sheets['Sheet1']
        book_name = sht1.range('A1').value
        sht1.range('O6').value = "制造分部\nMF"

        current_date = datetime.now().strftime('%Y%m%d')

        df1_1 = get_data_from_range(sht1, 'A6:S6', 1, 19)
        df1_2 = get_data_from_range(sht1, 'W6:Y6', 23, 25)
        df1_3 = get_data_from_range(sht1, 'T8:U8', 20, 21)
        df1 = pd.concat([df1_1, df1_2, df1_3], axis=1)

        columns_to_drop = ['工艺\nMET', '外控\nSQC', '采购\nPROC', '运维\nO&M', '项目\nPM', '工程\nENG',
                           '售后质量', 'Engineer', '质量-外控', '售后质量', '采购', '运维', '项目', '工程']

        data = []
        max_column = sht1.used_range.last_cell.column
        for col_index in range(ord('Z') - ord('A'), max_column + 1):
            col_letter = get_column_letter(col_index + 1)
            cell = sht1[f'{col_letter}8']
            start_row_num, start_col_num, end_row_num, end_col_num = get_merge_info(cell)
            start_address = f"{get_column_letter(start_col_num)}{start_row_num}"
            end_address = f"{get_column_letter(end_col_num)}{end_row_num}"
            merge_count = (end_col_num - start_col_num + 1) * (end_row_num - start_row_num + 1) if cell.api.MergeCells else 1
            merge_context = cell.value if cell.api.MergeCells else ''
            data.append({
                'Start_Address': start_address,
                'End_Address': end_address,
                'Start_Column': start_col_num,
                'End_Column': end_col_num,
                'Merged_Cell_Count': merge_count,
                'Merge_Context': merge_context
            })

        df_info = pd.DataFrame(data)
        df_info.replace('', np.nan, inplace=True)
        df_info.drop_duplicates(inplace=True)
        df_info.dropna(subset=['Merge_Context'], inplace=True)

        excel_file_name = f"{book_name}_SingleCar_{current_date}.xlsx"

        with pd.ExcelWriter(excel_file_name, engine='openpyxl') as writer:
            workbook = writer.book
            for index, row in df_info.iterrows():
                start_col_num = int(row['Start_Column'])
                end_col_num = int(row['End_Column'])
                merge_context = row['Merge_Context']
                merged_df, sum_CN, sum_MF_CN, sum_MF_OPEN_CN = process_sheet(sht1, start_col_num, end_col_num, merge_context)

                merged_df.to_excel(writer, sheet_name=f"{merge_context}", startrow=2, startcol=0, index=False)

                activesheet = writer.sheets[f"{merge_context}"]
                activesheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=activesheet.max_column)
                activesheet.cell(row=1, column=1).value = f"{merge_context}车辆变更状态"
                activesheet.cell(row=2, column=1).value = "变更总数："
                activesheet.cell(row=2, column=2).value = sum_CN
                activesheet.cell(row=2, column=3).value = "车间变更数："
                activesheet.cell(row=2, column=4).value = sum_MF_CN
                activesheet.cell(row=2, column=5).value = "未关闭："
                activesheet.cell(row=2, column=6).value = sum_MF_OPEN_CN

                for col_index in range(1, 7):
                    activesheet.column_dimensions[get_column_letter(col_index)].width = 16.5

                fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                font = Font(name='黑体', bold=True, size=14)
                alignment = Alignment(horizontal="center", vertical="center")

                apply_format(activesheet.cell(row=1, column=1), font_size=16)
                for cell in [activesheet.cell(row=2, column=1), activesheet.cell(row=2, column=2),
                             activesheet.cell(row=2, column=3), activesheet.cell(row=2, column=4),
                             activesheet.cell(row=2, column=5), activesheet.cell(row=2, column=6)]:
                    apply_format(cell)

except Exception as e:
    print(f"写入文件时发生错误: {e}")
finally:
    print(f"数据处理结束。")