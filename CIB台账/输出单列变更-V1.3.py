# -*- coding: utf-8 -*-

import pandas as pd
import time
from datetime import datetime
import numpy as np
import tkinter as tk
from tkinter import filedialog
import xlwings as xw
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font,Border,Side
from tqdm import tqdm

def select_excel_file():
    """选择Excel文件的对话框"""
    return filedialog.askopenfilename(title="选择Excel文件", filetypes=[("Excel Files", "*.xlsx;*.xls")])
def get_merge_info(cell):
    """获取合并单元格的信息"""
    if cell.api.MergeCells:
        merge_area = cell.api.MergeArea
        return (
            merge_area.Cells(1).Row,  # 起始行号
            merge_area.Cells(1).Column,  # 起始列号
            merge_area.Cells(merge_area.Rows.Count, 1).Row,  # 结束行号
            merge_area.Cells(1, merge_area.Columns.Count).Column  # 结束列号
        )
    else:
        return cell.row, cell.column, cell.row, cell.column


# 定义函数获取数据
def get_data_from_range(sheet, header_range, start_col_num, end_col_num):
    start_col = xw.utils.col_name(start_col_num)
    end_col = xw.utils.col_name(end_col_num)
    headers = sheet.range(header_range).value
    start_row = 10
    last_row = sheet.cells.last_cell.row
    end_row = sheet.range(f'{end_col}{last_row}').end('up').row
    data_range = sheet.range(f'{start_col}{start_row}:{end_col}{end_row}')
    data = data_range.value
    df = pd.DataFrame(data, columns=headers[:len(data[0])])
    df.dropna(how='all', inplace=True)
    return df

def get_unfinished_data(str1, df_OPEN):
    # 1. 找出所有包含str1的列名
    index_lst = [col for col in df_OPEN.columns if str1 in col]

    # 2. 筛选出含有“未完成”的行
    # 首先，构造一个布尔索引，用于标记哪些行在index_lst中的任一列包含“未完成”
    mask = df_OPEN.apply(lambda row: any('未完成' in str(row[col]) for col in index_lst), axis=1)

    # 应用筛选
    df_OPEN_filtered = df_OPEN[mask]
    sum_filter_OPEN = df_OPEN_filtered.shape[0]

    return sum_filter_OPEN
    

# 提取合并单元格信息
def extract_merge_info(sheet, start_row, start_col):
    """提取合并单元格的信息，并过滤掉Merge_Context为空的记录"""
    max_column = sheet.used_range.last_cell.column
    data = []
    for col_index in range(ord(start_col) - ord('A'), max_column + 1):
        col_letter = get_column_letter(col_index + 1)
        cell = sheet[f'{col_letter}{start_row}']
        try:
            start_row_num, start_col_num, end_row_num, end_col_num = get_merge_info(cell)
            start_address = f"{get_column_letter(start_col_num)}{start_row_num}"
            end_address = f"{get_column_letter(end_col_num)}{end_row_num}"
            merge_count = (end_col_num - start_col_num + 1) * (end_row_num - start_row_num + 1) if cell.api.MergeCells else 1
            merge_context = cell.value if cell.api.MergeCells else ''
            # 直接在这里过滤掉merge_context为空的记录
            if merge_context:
                data.append({
                    'Start_Address': start_address,
                    'End_Address': end_address,
                    'Start_Column': start_col_num,
                    'End_Column': end_col_num,
                    'Merged_Cell_Count': merge_count,
                    'Merge_Context': merge_context
                })
        except Exception as e:
            print(f"处理列 {col_letter} 时发生错误: {e}")
    return data
# 添加边框到B1:L4单元格
def add_borders(sheet, start_row=1, start_col=2, end_row=4, end_col=12):
    """
    为指定单元格范围添加边框。
    :param sheet: 工作表对象。
    :param start_row: 起始行号。
    :param start_col: 起始列号。
    :param end_row: 结束行号。
    :param end_col: 结束列号。
    """
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = thin_border
def apply_format(cell, font_name='等线', font_size=14, font_color='FF0000', is_bold=True, border=True, align_center=True):
    """应用格式到单元格"""
    font = Font(name=font_name, color=font_color, bold=is_bold, size=font_size)
    cell.font = font
    if border:
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        cell.border = thin_border
    if align_center:
        alignment = Alignment(horizontal='center', vertical='center')
        cell.alignment = alignment


# 选择文件
file_path = select_excel_file()
if not file_path:
    print("没有选择文件。")



sheet_name = 'Sheet1'

# 打开Excel文件并激活工作表
wb = xw.Book(file_path)
sht1 = wb.sheets[sheet_name]
book_name=sht1.range('A1').value
sht1.range('O6').value="计划\nPlan" #修改原表中的制造分部df1中标题防止出现重复列

# 获取Dataframe1区域的数据
df1_1 = get_data_from_range(sht1, 'A6:S6', 1, 19)
df1_2 = get_data_from_range(sht1, 'W6:Y6', 23, 25)
df1_3 = get_data_from_range(sht1, 'T8:U8', 20, 21)
df1 = pd.concat([df1_1, df1_2,df1_3], axis=1)

# 移除df1中不必要的列
# 定义需要移除的列名列表
columns_to_drop = ['工艺\nMET', '变更通知编号\nECN','外控\nSQC', '内控\nQC','采购\nPROC', '运维\nO&M', '项目\nPM', '工程\nENG', 
                       '售后质量', '基线范围\nScope of influence','供应商整改范围\nScope of Supplier','Engineer', '质量-外控','质量-内控','售后质量', '运维', '项目', '工程']
# 移除指定的列
df1.drop(columns=columns_to_drop, errors='ignore', inplace=True)


# 获取当前日期并格式化为年月日形式
current_date = datetime.now().strftime('%Y%m%d %H-%M')

# 使用日期作为后缀创建新工作簿的名称
excel_file_name = f"{book_name}_TS_{current_date}.xlsx"

print("开始读取数据,不要关闭工作表，请耐心等待。。。")
data_TS = extract_merge_info(sht1, 6, 'Z')
data_Car = extract_merge_info(sht1, 8, 'Z')

df_TS_info = pd.DataFrame(data_TS)  # 提取TS信息,车辆列的合并单元格信息
df_Car_info = pd.DataFrame(data_Car) #提取Car信息,单个车的合并单元格信息

# print(df_TS_info)
# print(df_Car_info)
#遍历 df_Car_info,修改列名，增加车辆编号

for index, row in df_Car_info.iterrows():
    start_address = f"{get_column_letter(row['Start_Column'])}9"
    end_address = f"{get_column_letter(row['End_Column'])}9"
    car_NO= row['Merge_Context']
    for i in sht1.range(f"{start_address}: {end_address}"):
        # 判断range(i).value的值是否已经包含car_NO
        if car_NO not in i.value:
            sht1.range(i).value = f"{car_NO}\n{sht1.range(i).value}"




total_sheets = len(df_TS_info)
with tqdm(total=total_sheets, desc="Writing Sheets:", unit="sheets") as sheets_pbar:

    
    with pd.ExcelWriter(excel_file_name, engine='openpyxl') as writer:
        #遍历每列车工作表，并进行写入
        for index, row in df_TS_info.iterrows():
            start_col_num = int(row['Start_Column'])
            end_col_num = int(row['End_Column'])
            merge_context = row['Merge_Context']
            header_range = f"{get_column_letter(start_col_num)}9:{get_column_letter(end_col_num)}9"

            df_i = get_data_from_range(sht1, header_range, start_col_num, end_col_num)
            #获取df_i 标题，最后一个值为空值，需要进行修改，内容为 单元格({get_column_letter(end_col_num)}7)的值
                # 打印df_i的标题（假设df_i的列名为其标题）
            # print("Current Headers of df_i:")
            # print(df_i.columns)


            last_header = df_i.columns[-1]
            if pd.isnull(last_header):  # 检查是否为空值
                new_header = sht1.range(f"{get_column_letter(end_col_num)}7").value
                df_i.rename(columns={last_header: new_header}, inplace=True)  # 修改列名
                # print(f"Modified the last header to: {new_header}")
            else:
                print("The last header was not empty, no modification needed.")
            

            merged_df = pd.concat([df1.reset_index(drop=True), df_i.reset_index(drop=True)], axis=1)
            #把列名中包含 '售后质量', 'Engineer', '质量-外控', '售后质量', '采购', '运维', '项目', '工程' 的列清除
            columns_to_remove = [col for col in merged_df.columns if any(substring in col for substring in columns_to_drop)]
            merged_df.drop(columns=columns_to_remove, inplace=True, errors='ignore')
            
            # print(merged_df.columns)
            
            #将["列汇总"]这种的空值行清除                
            merged_df = merged_df.dropna(subset=["列汇总"])

            # 计算不含标题的实际数据行数，并命名变量为 sum_CN（表示该项目所有的变更总数）
            sum_CN = merged_df.shape[0]
            

            # 筛选条件：筛选出“列汇总”中含有“未完成”的行，并命名变量为 df_OPEN（表示该项目的未完成变更）
            df_OPEN = merged_df[merged_df['列汇总'].str.contains('未完成')]
            sum_OPEN = df_OPEN.shape[0]

            # # 1. 找出所有包含“生产”的列名
            # index_lst = [col for col in df_OPEN.columns if '生产' in col]

            # # 2. 筛选出含有“未完成”的行
            # # 首先，构造一个布尔索引，用于标记哪些行在index_lst中的任一列包含“未完成”
            # mask = df_OPEN.apply(lambda row: any('未完成' in str(row[col]) for col in index_lst), axis=1)

            # # 应用筛选
            # df_OPEN_filtered = df_OPEN[mask]

            sum_MF_OPEN_CN = get_unfinished_data('生产',df_OPEN)
            sum_OT_OPEN_CN = get_unfinished_data('外协',df_OPEN)
            sum_PL_OPEN_CN = get_unfinished_data('制造分部',df_OPEN)


        
            
            df_OPEN.to_excel(writer, sheet_name=f"{merge_context}", startrow=4, startcol=0, index=False)


            activesheet = writer.sheets[f"{merge_context}"]

            # 合并单元格
            activesheet.merge_cells(start_row=1, start_column=7, end_row=4, end_column=12)

            # 设置合并单元格的值
            activesheet.cell(row=1, column=7).value = f"{merge_context}车辆变更状态"
            activesheet.cell(row=1, column=2).value = "变更总数"
            activesheet.cell(row=1, column=3).value = sum_CN
            activesheet.cell(row=1, column=4).value = "未关闭"
            activesheet.cell(row=2, column=4).value = "车间"
            activesheet.cell(row=3, column=4).value = "外协"
            activesheet.cell(row=4, column=4).value = "计划"
            
                      
            activesheet.cell(row=1, column=5).value = sum_OPEN
            activesheet.cell(row=2,column=5).value = sum_MF_OPEN_CN
            activesheet.cell(row=3, column=5).value = sum_OT_OPEN_CN
            activesheet.cell(row=4, column=5).value = sum_PL_OPEN_CN



            #调整列宽
            activesheet.column_dimensions['B'].width = 16
            activesheet.column_dimensions['C'].width = 20



            # 创建自定义的填充样式
            fill = PatternFill(start_color="ADD8E6",  # 浅蓝色的RGB代码
                                end_color="ADD8E6",
                                fill_type="solid")
            
            # 设置字体为黑体
            font = Font(name='黑体', bold=True, size=14)  # 注意：'SimHei'一般用于中文黑体，英文环境下可能需要其他字体名称
            
            # 设置单元格内容居中
            alignment = Alignment(horizontal="center", vertical="center")

            # 应用样式
            activesheet.cell(row=1, column=7).fill = fill
            activesheet.cell(row=1, column=7).font = font
            activesheet.cell(row=1, column=7).alignment = alignment

            # 应用格式到指定的单元格
            apply_format(activesheet.cell(row=1, column=7), font_size=20)  # 标题行可能需要更大字体
            apply_format(activesheet.cell(row=1, column=2))
            apply_format(activesheet.cell(row=1, column=3))
            apply_format(activesheet.cell(row=1, column=4))
            apply_format(activesheet.cell(row=1, column=5))
            apply_format(activesheet.cell(row=2, column=4))
            apply_format(activesheet.cell(row=2, column=5))
            apply_format(activesheet.cell(row=3, column=4))
            apply_format(activesheet.cell(row=3, column=5))
            apply_format(activesheet.cell(row=4, column=4))
            apply_format(activesheet.cell(row=3, column=5))
            apply_format(activesheet.cell(row=4, column=5))

            add_borders(activesheet)



            # 设置第3行为标题行，并开启自动换行
            for col in activesheet.iter_cols(min_col=1, max_col=activesheet.max_column, min_row=5, max_row=5):
                for cell in col:
                    cell.font = Font(name='等线', size=11, bold=True)
                    cell.alignment = Alignment(wrapText=True, vertical='center', horizontal='center')
                    # cell.alignment = openpyxl.styles.Alignment(wrap_text=True)  # 启用自动换行

            # 调整第五行的行高
            activesheet.row_dimensions[5].height = 36
            # 确定数据起始行和结束行，以及列范围
            start_row = 5  # 数据起始行
            end_row = activesheet.max_row  # 数据结束行（假设数据连续到最后一行）
            column_letters = 'D'  # 假设数据从D列开始
            end_column = get_column_letter(activesheet.max_column)  # 获取最大列字母

            # 设置筛选范围，从第三行开始到工作表的最后一行和列
            data_range = f"{column_letters}{start_row}:{end_column}{end_row}"

            # 启用筛选
            activesheet.auto_filter.ref = data_range

                      
            
            
            sheets_pbar.update(1)
    


print(f"数据已成功写入到 {excel_file_name}")
# 关闭原工作簿，根据需要可取消注释
wb.save()
wb.app.quit()



# ...（此处是你的程序代码）

# 程序完成操作后，等待5秒
print("程序执行完毕，即将退出...")
time.sleep(5)  # 暂停5秒

# 程序退出
exit(0)  # 可选，正常退出程序，0表示成功