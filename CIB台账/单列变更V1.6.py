# -*- coding: utf-8 -*-
import pandas as pd
import numpy as np
import re
import time
from datetime import datetime
from tkinter import filedialog
import xlwings as xw
# import openpyxl
import os
import json
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font,Border,Side
from tqdm import tqdm
# from openpyxl.utils.cell import coordinate_from_string
# from openpyxl.utils import column_index_from_string
from openpyxl import Workbook
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from openpyxl.utils.dataframe import dataframe_to_rows
import seaborn as sns


def select_excel_file():
    """选择Excel文件的对话框"""
    return filedialog.askopenfilename(title="选择Excel文件", filetypes=[("Excel Files", "*.xlsx;*.xls")])

note ="""
    ┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄
                                    使用前注意事项
    0.本程序支持对Office Excel文件格式的变更处理,将变更汇总表按单列进行拆分

    1.被处理的CIB台帐请确保是从QMIS系统中自动导出的台账格式！，并且清除所有筛选条件

    2.本程序支持主要依据CIB导出的台账进行案列拆分，拆分后对未完成的变更的台账会自动保存到当前目录下
    
    3.如果拆分后发现在同一时间进行两次及以上，使用时会出现报错，等待一分钟后重新运行即可

    4.本程序运行中会对原文件中标题内容有部分改动，未对数据进行修改
    
    5.程序运行中会弹出一个窗口，请勿关闭窗口，否则会报错，运行时间取决于项目车辆数量，如果项目车辆数量较多，请耐心等待
    
    6.程序运行完成后将产生一个history.json文件，保存了每个车的变更信息，请勿删除或者修改文件，确保该文件与本程序放在同一文件夹下

    7.程序运行完成后会产生一个图示化照片，显示最近两天的每个车每个部门完成/未完成的数量

    8.本程序为V1.6版本，后续版本会更新，敬请关注

    ┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄  
    """
print (note)
ss = input("====确认后，按回车继续====\n")

print("====请选择需要处理的CIB台账====")

file_path = select_excel_file()
if not file_path:
    print("没有选择文件。")
sheet_name = 'Sheet1'


# 打开Excel文件并激活工作表
wb = xw.Book(file_path)
sht1 = wb.sheets[sheet_name]
book_name=sht1.range('A1').value
sht1.range('O6').value="计划\nPlan" #修改原表中的制造分部df1中标题防止出现重复列
 # 定义需要移除的列名列表
columns_to_drop = ['工艺\nMET', '变更通知编号\nECN','外控\nSQC', '内控\nQC','采购\nPROC', '运维\nO&M', '项目\nPM', '工程\nENG', 
                        '售后质量', '基线范围\nScope of influence','供应商整改范围\nScope of Supplier','Engineer', '质量-外控','质量-内控','售后质量', '运维', '项目', '工程']

# 获取当前日期并格式化为年月日形式
current_datetime = datetime.now().strftime('%Y%m%d %H-%M')
# 使用日期作为后缀创建新工作簿的名称
excel_file_name = f"{book_name}_TS_{current_datetime}.xlsx"
# 设置matplotlib字体以支持中文
plt.rcParams['font.sans-serif'] = ['SimHei']  # 使用微软雅黑字体支持中文
plt.rcParams['axes.unicode_minus'] = False  # 正确显示负号

def main():
    # 假设 sht1 是一个有效的表格或工作表对象，book_name 是项目的名称
    df_TS_info, df_Car_info, train_consist, project_ID = load_or_calculate_data(sht1, book_name)


    # 调用函数,构造出车辆结构
    # nested_dict = build_project_car_dict(project_ID,df_TS_info, df_Car_info, train_consist)
    
    total_sheets = len(df_TS_info)

    df1_1 = get_data_from_range(sht1, 'A6:S6', 1, 19)
    df1_2 = get_data_from_range(sht1, 'W6:Y6', 23, 25)
    df1_3 = get_data_from_range(sht1, 'T8:U8', 20, 21)
    df1 = pd.concat([df1_1, df1_2,df1_3], axis=1)

    # 移除df1中不必要的列
    df1.drop(columns=columns_to_drop, errors='ignore', inplace=True)
    #使用正则表达式将df1标题中\n之后的英文去除
    df1.columns = [re.sub(r'\n[a-zA-Z ]*', '', name) for name in df1.columns]

    #遍历 df_Car_info,修改列名，增加车辆编号
    for index, row in df_Car_info.iterrows():
        start_address = f"{get_column_letter(row['Start_Column'])}9"
        end_address = f"{get_column_letter(row['End_Column'])}9"
        car_NO= row['Merge_Context']
        for i in sht1.range(f"{start_address}: {end_address}"):
            # 判断range(i).value的值是否已经包含car_NO
            if car_NO not in i.value:
                sht1.range(i).value = f"{car_NO}\n{sht1.range(i).value}"

    # wb1 = Workbook()
    # ws = wb1.active
    # ws.title = "Sheet1"  # 设置工作表的标题
    with tqdm(total=total_sheets, desc="Writing Sheets:", unit="sheets") as sheets_pbar:

        with pd.ExcelWriter(excel_file_name, engine='openpyxl',mode='w') as writer:
            
            # 读取现有history.json文件中的数据
            try:
                with open('history.json', 'r', encoding='utf-8') as file:
                    existing_data = json.load(file)
            except FileNotFoundError:
                existing_data = {}

            #遍历每列车工作表，并进行写入
            for index, row in df_TS_info.iterrows():
                start_col_num = int(row['Start_Column'])
                end_col_num = int(row['End_Column'])
                merge_context = row['Merge_Context']
                header_range = f"{get_column_letter(start_col_num)}9:{get_column_letter(end_col_num)}9"

                df_i = get_data_from_range(sht1, header_range, start_col_num, end_col_num)
                
                last_header = df_i.columns[-1]
                if pd.isnull(last_header):  # 检查是否为空值
                    new_header = sht1.range(f"{get_column_letter(end_col_num)}7").value
                    df_i.rename(columns={last_header: new_header}, inplace=True)  # 修改列名
                    # print(f"Modified the last header to: {new_header}")
                else:
                    print("The last header was not empty, no modification needed.")
            
                
                merged_df = pd.concat([df1.reset_index(drop=True), df_i.reset_index(drop=True)], axis=1)
                #把列名中包含 '售后质量', 'Engineer', '质量-外控', '售后质量', '采购', '运维', '项目', '工程' 的列清除
                columns_to_remove = [col for col in merged_df.columns if any(substring in col for substring in columns_to_drop)]
                merged_df.drop(columns=columns_to_remove, inplace=True, errors='ignore')
                
                               
                #将["列汇总"]这种的空值行清除                
                merged_df = merged_df.dropna(subset=["列汇总"])

                # 计算不含标题的实际数据行数，并命名变量为 sum_CN（表示该项目所有的变更总数）
                sum_CN = merged_df.shape[0]
                
                # 筛选条件：筛选出“列汇总”中含有“未完成”的行，并命名变量为 df_OPEN（表示该项目的未完成变更）
                df_OPEN = merged_df[merged_df['列汇总'].str.contains('未完成')]
                sum_OPEN = df_OPEN.shape[0]

                sum_MF_OPEN_CN = get_unfinished_data('生产',df_OPEN)
                sum_OT_OPEN_CN = get_unfinished_data('外协',df_OPEN)
                sum_PL_OPEN_CN = get_unfinished_data('制造分部',df_OPEN)

                # 假设merged_df是已经定义好的DataFrame
                car_mf_log  = count_completed_uncompleted(df_i, '生产')
                car_OT_log  = count_completed_uncompleted(df_i, '外协')
                car_PROC_log = count_completed_uncompleted(df_i, '采购')

                # 构建history_log字典
                current_date = datetime.now().strftime('%Y年%m月%d日')
                book_name_history = f"{book_name}history"                    
                Project_ID = book_name_history
                TS = merge_context

                if Project_ID not in existing_data:
                    existing_data[Project_ID] = {}
                if TS not in existing_data[Project_ID]:
                    existing_data[Project_ID][TS] = {}
                if current_date not in existing_data[Project_ID][TS]:
                    existing_data[Project_ID][TS][current_date] = {}
                
                existing_data[Project_ID][TS][current_date]["car_MF_log"] = car_mf_log
                existing_data[Project_ID][TS][current_date]["car_OT_log"] = car_OT_log
                existing_data[Project_ID][TS][current_date]["car_PROC_log"] = car_PROC_log
                # existing_data[Project_ID][TS][current_date]["Sum_MF_OPEN_CN"] = sum_MF_OPEN_CN
                # existing_data[Project_ID][TS][current_date]["Sum_OT_OPEN_CN"] = sum_OT_OPEN_CN
                # existing_data[Project_ID][TS][current_date]["Sum_PL_OPEN_CN"] = sum_PL_OPEN_CN
                             
                df_OPEN.to_excel(writer, sheet_name=f"{merge_context}", startrow=4, startcol=0, index=False)
                activesheet = writer.sheets[f"{merge_context}"]

                format_worksheet(activesheet, merge_context, sum_CN, sum_OPEN, sum_MF_OPEN_CN, sum_OT_OPEN_CN, sum_PL_OPEN_CN)

                sheets_pbar.update(1)

    with open('history.json', 'w', encoding='utf-8') as file:
                json.dump(existing_data, file, ensure_ascii=False, indent=4)
                # print("History data has been updated.")     
    
    print(f"数据已成功写入到 {excel_file_name}\n接下来，更新history.json文件")
    # 关闭原工作簿，根据需要可取消注释
    wb.save()
    wb.app.quit()


    # 对history.json的存储数据进行删除记录并进行扁平化转化为DataFrame.
    with open('history.json', 'r', encoding='utf-8') as file:
        data = json.load(file)
        for ts, ts_data in data[f"{book_name}history"].items():

            # 处理历史记录,将date数据多余的数据进行删除
            data[f"{book_name}"][ts] = process_history(data[f"{book_name}history"][ts])
        
        history_data = data[f"{book_name}history"]
        
        # 创建一个空列表，用于存储扁平化的数据
        flat_history = []

        # 遍历 history 数据
        for ts, ts_data in history_data.items():            
            for date, date_data in ts_data.items():
                for car_type, car_type_data in date_data.items():
                    for car, car_data in car_type_data.items():
                        
                        completed_count = car_data["已完成数量"]
                        remaining_count = car_data["未完成数量"]
                        # car_dep=[line.split("\n") for line in car.strip().splitlines()]
                        data1, data2 = car.split('\n')
                        flat_history.append({
                            "车号": data1,
                            "部门": data2,
                            "日期": date,
                            "已完成数量": completed_count,
                            "未完成数量": remaining_count
                        })

        # 将扁平化的数据转换为 DataFrame
        df_history = pd.DataFrame(flat_history)
        # 对 DataFrame 进行排序
        df_history = df_history.sort_values(by=['车号', '部门', '日期'], ascending=[True, True, True])

    # 保存更新后的 Json 数据

    with open('history.json', 'w', encoding='utf-8') as file:
        json.dump(data, file, ensure_ascii=False, indent=4)
        print(f"history.json 文件更新成功.\n接下来，将历史数据更新到{excel_file_name}文件中")
     
    append_df_to_excel(df_history, excel_file_name, sheet_name=f"history")
    print(f"数据已成功写入到 {excel_file_name}\n接下来，将创建可视化图")          

    df_history['日期'] = pd.to_datetime(df_history['日期'], format="%Y年%m月%d日", errors='coerce')
    df_povit = process_data(df_history)
    plot_data(df_povit, train_consist, colors=[ 'lightcyan', 'thistle',
                                                 'palegoldenrod', 'slategrey', 'thistle', 'plum',
                                                'lightseagreen', 'skyblue', 'palevioletred', 'lightsteelblue',
                                                'lightgrey','lightblue','lightsalmon', 'lightcoral', 'lightgreen'])    



    


    # 程序完成操作后，等待5秒
    print("程序执行完毕，即将退出...")
    time.sleep(5)  # 暂停5秒
    

    input("按回车键退出程序...")
    # 程序退出
    exit(0)  # 可选，正常退出程序，0表示成功




def append_df_to_excel(df, excel_file_name, sheet_name='New Sheet'):
    """
    将DataFrame追加到Excel工作簿的新工作表中。
    :param df: pandas.DataFrame，要写入的DataFrame
    :param excel_file_name: str，Excel文件的路径
    :param sheet_name: str，新工作表的名称
    """
    # 加载现有的Excel工作簿
    workbook = load_workbook(excel_file_name)

    # 创建一个新的工作表
    if sheet_name not in workbook.sheetnames:
        worksheet = workbook.create_sheet(sheet_name)
    else:
        worksheet = workbook[sheet_name]

    # 将DataFrame的列名和值写入工作表
    for row in dataframe_to_rows(df, index=False, header=True):
        worksheet.append(row)

    # 确保新工作表位于所有现有工作表之前
    if sheet_name != workbook.sheetnames[0]:
        # 直接操作_sheets属性
        sheets = workbook._sheets
        new_sheet = workbook[sheet_name]
        sheets.remove(new_sheet)
        sheets.insert(0, new_sheet)

    # 保存工作簿
    workbook.save(excel_file_name)
    workbook.close()
    print(f"历史数据已更新到 {excel_file_name}")
    return

def flatten_json(nested_json, prefix=''):
    out = {}
    for key, value in nested_json.items():
        if isinstance(value, dict):
            out.update(flatten_json(value, prefix + key + '_'))
        elif isinstance(value, list):
            for i, item in enumerate(value):
                if isinstance(item, dict):
                    out.update(flatten_json(item, prefix + key + '_' + str(i) + '_'))
                else:
                    out[prefix + key + '_' + str(i)] = item
        else:
            out[prefix + key] = value
    return out

def process_history(history_data):
   # 获取历史记录的键（即日期）
    dates = sorted(history_data.keys())

    while len(dates) > 30:
        oldest_date = dates.pop(0)
        del history_data[oldest_date]
        print(f"历史内容过多，历史记录中{oldest_date}的记录被删除")    
    return history_data


def format_worksheet(activesheet, merge_context, sum_CN, sum_OPEN, sum_MF_OPEN_CN, sum_OT_OPEN_CN, sum_PL_OPEN_CN):
    """
    格式化TS工作表，包括合并单元格、设置单元格值、调整列宽、应用样式、设置标题行、调整行高等操作。
    
    参数:
        activesheet (Worksheet): 需要格式化的Excel工作表对象。
        merge_context (str): 合并单元格后显示的文本前缀。
        sum_CN (int): 变更总数。
        sum_OPEN (int): 未关闭的变更数量总和。
        sum_MF_OPEN_CN (int): 车间未关闭的变更数量。
        sum_OT_OPEN_CN (int): 外协未关闭的变更数量。
        sum_PL_OPEN_CN (int): 计划未关闭的变更数量。
        
    返回:
        None: 直接修改传入的工作表对象。
    """

    # 合并单元格
    activesheet.merge_cells(start_row=1, start_column=7, end_row=4, end_column=12)

    # 设置合并单元格的值
    activesheet.cell(row=1, column=7).value = f"{merge_context}车辆变更状态"

    # 设置其他单元格的值
    activesheet.cell(row=1, column=2).value = "变更总数"
    activesheet.cell(row=1, column=3).value = sum_CN
    activesheet.cell(row=1, column=4).value = "未关闭"
    activesheet.cell(row=2, column=4).value = "车间"
    activesheet.cell(row=3, column=4).value = "外协"
    activesheet.cell(row=4, column=4).value = "计划"
    activesheet.cell(row=1, column=5).value = sum_OPEN
    activesheet.cell(row=2, column=5).value = sum_MF_OPEN_CN
    activesheet.cell(row=3, column=5).value = sum_OT_OPEN_CN
    activesheet.cell(row=4, column=5).value = sum_PL_OPEN_CN

    # 调整列宽
    activesheet.column_dimensions['B'].width = 16
    activesheet.column_dimensions['C'].width = 20

    # 创建自定义的填充样式
    fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    # 设置字体为黑体
    font = Font(name='黑体', bold=True, size=14)

    # 设置单元格内容居中
    alignment = Alignment(horizontal="center", vertical="center")

    # 应用样式
    activesheet.cell(row=1, column=7).fill = fill
    activesheet.cell(row=1, column=7).font = font
    activesheet.cell(row=1, column=7).alignment = alignment

    # 应用格式到指定的单元格
    apply_format(activesheet.cell(row=1, column=7), font_size=20)
    for row in range(1, 5):
        for col in range(2, 6):
            apply_format(activesheet.cell(row=row, column=col))

    add_borders(activesheet)

    # 设置第3行为标题行，并开启自动换行
    for col in activesheet.iter_cols(min_col=1, max_col=activesheet.max_column, min_row=5, max_row=5):
        for cell in col:
            cell.font = Font(name='等线', size=11, bold=True)
            cell.alignment = Alignment(wrapText=True, vertical='center', horizontal='center')

    # 调整第五行的行高
    activesheet.row_dimensions[5].height = 36

    # 设置筛选范围
    start_row = 5
    end_row = activesheet.max_row
    column_letters = 'D'
    end_column = get_column_letter(activesheet.max_column)
    data_range = f"{column_letters}{start_row}:{end_column}{end_row}"

    # 启用筛选
    activesheet.auto_filter.ref = data_range

def load_or_calculate_data(sht1, book_name):
    # 定义历史数据文件路径
    history_path = "history.json"

    # 初始化history_data为一个空字典
    history_data = {}

    # 尝试从文件加载历史数据
    if os.path.exists(history_path):
        try:
            with open(history_path, 'r',encoding='utf-8') as file:
                history_data = json.load(file)
                
            # 检查是否有与book_name匹配的project_ID
            if book_name in history_data:
                df_TS_info = pd.DataFrame(history_data[book_name]['df_TS_info'])
                df_Car_info = pd.DataFrame(history_data[book_name]['df_Car_info'])
                train_consist = history_data[book_name]['train_consist']
                print(f"Data for project '{book_name}' loaded from history.json.")
                return df_TS_info, df_Car_info, train_consist, book_name
            
            else:
                print(f"No data found for project '{book_name}' in history.json. Calculating new data...")
        
        except Exception as e:
            print(f"Error loading data from history.json: {e}")

    # 如果文件不存在或没有找到匹配的book_name，执行原始代码逻辑
    data_TS = extract_merge_info(sht1, 6, 'Z')
    data_Car = extract_merge_info(sht1, 8, 'Z')

    df_TS_info = pd.DataFrame(data_TS)
    df_Car_info = pd.DataFrame(data_Car)
    

    try:
        train_consist = (df_TS_info['Merged_Cell_Count'][0] - 1) / df_Car_info['Merged_Cell_Count'][0]
    except Exception as e:
        print(f"Error calculating train_consist: {e}")
        train_consist = None

    # 更新history.json文件
    try:
        # 添加当前book_name的数据
        history_data[book_name] = {
            'df_TS_info': df_TS_info.to_dict(orient='records'),
            'df_Car_info': df_Car_info.to_dict(orient='records'),
            'train_consist': train_consist,
            'project_ID': book_name
        }

        with open(history_path, 'w',encoding='utf-8') as file:
            json.dump(history_data, file, ensure_ascii=False)
        print(f"Data for project '{book_name}' saved to history.json.")
    except Exception as e:
        print(f"Error saving data to history.json: {e}")

    return df_TS_info, df_Car_info, train_consist, book_name


def get_merge_info(cell):
    """获取合并单元格的信息"""
    if cell.api.MergeCells:
        merge_area = cell.api.MergeArea
        return (
            merge_area.Cells(1).Row,  # 起始行号
            merge_area.Cells(1).Column,  # 起始列号
            merge_area.Cells(merge_area.Rows.Count, 1).Row,  # 结束行号
            merge_area.Cells(1, merge_area.Columns.Count).Column  # 结束列号
        )
    else:
        return cell.row, cell.column, cell.row, cell.column
    
# 定义函数获取数据
def get_data_from_range(sheet, header_range, start_col_num, end_col_num):
    start_col = xw.utils.col_name(start_col_num)
    end_col = xw.utils.col_name(end_col_num)
    headers = sheet.range(header_range).value
    start_row = 10
    last_row = sheet.cells.last_cell.row
    end_row = sheet.range(f'{end_col}{last_row}').end('up').row
    data_range = sheet.range(f'{start_col}{start_row}:{end_col}{end_row}')
    data = data_range.value
    df = pd.DataFrame(data, columns=headers[:len(data[0])])
    df.dropna(how='all', inplace=True)
    return df

def get_unfinished_data(str1, df_OPEN):
    # 1. 找出所有包含str1的列名
    index_lst = [col for col in df_OPEN.columns if str1 in col]

    # 2. 筛选出含有“未完成”的行
    # 首先，构造一个布尔索引，用于标记哪些行在index_lst中的任一列包含“未完成”
    mask = df_OPEN.apply(lambda row: any('未完成' in str(row[col]) for col in index_lst), axis=1)

    # 应用筛选
    df_OPEN_filtered = df_OPEN[mask]
    sum_filter_OPEN = df_OPEN_filtered.shape[0]

    return sum_filter_OPEN


def add_borders(sheet, start_row=1, start_col=2, end_row=4, end_col=12):
    """
    为指定单元格范围添加边框。
    :param sheet: 工作表对象。
    :param start_row: 起始行号。
    :param start_col: 起始列号。
    :param end_row: 结束行号。
    :param end_col: 结束列号。
    """
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = thin_border



def apply_format(cell, font_name='等线', font_size=14, font_color='FF0000', is_bold=True, border=True, align_center=True):
    """应用格式到单元格"""
    font = Font(name=font_name, color=font_color, bold=is_bold, size=font_size)
    cell.font = font
    if border:
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        cell.border = thin_border
    if align_center:
        alignment = Alignment(horizontal='center', vertical='center')
        cell.alignment = alignment


# 提取合并单元格信息
def extract_merge_info(sheet, start_row, start_col):
    """提取合并单元格的信息，并过滤掉Merge_Context为空的记录"""
    max_column = sheet.used_range.last_cell.column
    data = []
    # data_lst = []
    for col_index in range(ord(start_col) - ord('A'), max_column + 1):
        col_letter = get_column_letter(col_index + 1)
        cell = sheet[f'{col_letter}{start_row}']
        try:
            start_row_num, start_col_num, end_row_num, end_col_num = get_merge_info(cell)
            start_address = f"{get_column_letter(start_col_num)}{start_row_num}"
            end_address = f"{get_column_letter(end_col_num)}{end_row_num}"
            merge_count = (end_col_num - start_col_num + 1) * (end_row_num - start_row_num + 1) if cell.api.MergeCells else 1
            merge_context = cell.value if cell.api.MergeCells else ''
            # 直接在这里过滤掉merge_context为空的记录
            if merge_context:
                # data_lst.append(merge_context)
                data.append({
                    'Start_Address': start_address,
                    'End_Address': end_address,
                    'Start_Column': start_col_num,
                    'End_Column': end_col_num,
                    'Merged_Cell_Count': merge_count,
                    'Merge_Context': merge_context
                })
    


        except Exception as e:
            print(f"处理列 {col_letter} 时发生错误: {e}")
            
        
    # print(f"data多级字典:{data_lst}")
    
    return data



def count_completed_uncompleted(df, prefix):
    """
    统计DataFrame中特定前缀列的已完成和未完成数量。
    
    :param df: 要处理的DataFrame。
    :param prefix: 列名前缀，用于筛选列。
    :return: 包含每列的已完成和未完成数量的字典。
    """
    mf_col_list = [col for col in df.columns if prefix in col]
    
    car_mf_log = {}
    #遍历mf_col_list，使用布尔索引计算已完成的数量，即排除含有"未完成"字符串和等于"/"的项
    for col in mf_col_list:
        completed_mask = ~(df[col].str.contains('未完成')) & (df[col] != '/')
        completed_count = int(completed_mask.sum())

        # 使用布尔索引计算未完成的数量，即包含"未完成"字符串的项
        unfinished_mask = df[col].str.contains('未完成')
        unfinished_count = int(unfinished_mask.sum())
        # 将当前列的已完成和未完成数量存入字典
        car_mf_log[col] = {
            '已完成数量': completed_count,
            '未完成数量': unfinished_count
                            }
                
    return car_mf_log

def process_data(df):
    try:
        # 使用pivot_table重塑数据
        df_pivot = df.pivot_table(
            index=['车号', '日期'],
            columns='部门',
            values=['已完成数量', '未完成数量'],
            aggfunc='sum'
        )

        # 将MultiIndex的第二层堆叠为长格式数据，使用新的stack()行为
        df_stacked = df_pivot.stack(future_stack=True).reset_index()
        
        # 将堆叠后的索引转换为列，保留所有索引级别
        df_stacked.columns = ['车号', '日期', '部门', '已完成数量', '未完成数量']

        # 检查'日期'列的数据类型，如果需要则转换为datetime类型
        if not np.issubdtype(df_stacked['日期'].dtype, np.datetime64):
            df_stacked['日期'] = pd.to_datetime(df_stacked['日期'], errors='coerce')
    except Exception as e:
        print(f"数据处理出错：{e}")
        raise

    return df_stacked

def plot_data(df_stacked, train_consist, colors):
    car_number_to_index = {car_number: int(index) for index, car_number in enumerate(df_stacked['车号'].unique())}
    df_stacked['车号_index'] = df_stacked['车号'].map(car_number_to_index)
    if train_consist<3:
        t=6
    else:
        t=train_consist

    unique_dates = df_stacked['日期'].unique()

    if len(unique_dates) > 1:
        latest_date = unique_dates[-1]
        previous_date = unique_dates[-2]
    else:
        # 处理只有一个日期的情况
        latest_date = unique_dates[0]
        previous_date = latest_date  # 或者设置为None，或者使用其他逻辑
        print(f"只有一个日期的历史数据，使用[{latest_date.strftime('%Y%m%d')}]作为最新日期")

    date_colors = {
        latest_date: 'orchid',  
        previous_date: 'lightgrey' 
    }

    num_cars = len(car_number_to_index)
    num_subplots = int(np.ceil(num_cars / t))
    fig, axs = plt.subplots(nrows=num_subplots, figsize=(10, 3*num_subplots))

    if not isinstance(axs, np.ndarray):
        axs = [axs]

    for i, ax in enumerate(axs):
        start_index = int(i * t)
        end_index = int(min((i + 1) * t, num_cars))
        subset_data = df_stacked[df_stacked['车号_index'].between(start_index, end_index-1)]
        if  len(unique_dates) > 1:
            ax.set_title(f"日期：{previous_date.strftime('%Y%m%d')}[浅灰色]和{latest_date.strftime('%Y%m%d')}[淡紫色]  车号 :{list(car_number_to_index.keys())[start_index]} -- {list(car_number_to_index.keys())[end_index-1]}")
        else:
            ax.set_title(f"日期：{latest_date.strftime('%Y%m%d')}[浅灰色]  车号 :{list(car_number_to_index.keys())[start_index]} -- {list(car_number_to_index.keys())[end_index-1]}")
        ax.set_ylabel('数量')

        x_positions = np.arange(start_index, end_index)

        for date, x_pos in zip([previous_date, latest_date], [x_positions - 0.2, x_positions + 0.2]):
            day_data = subset_data[subset_data['日期'] == date]

            bottom_values = np.zeros(len(x_pos))

            for department in df_stacked['部门'].unique():
                dept_data = day_data[day_data['部门'] == department]
                completed_counts = dept_data['已完成数量'].values
                uncompleted_counts = dept_data['未完成数量'].values
                total_counts = completed_counts + uncompleted_counts

                color_index = list(df_stacked['部门'].unique()).index(department)
                dept_color = colors[color_index]

                bars = ax.bar(x_pos, total_counts, bottom=bottom_values, width=0.4, align='center',
                              color=dept_color, edgecolor=date_colors[date], label=f"{department} ({date.strftime('%Y%m%d')})")

                bottom_values += total_counts

                for bar, cc, uc in zip(bars, completed_counts, uncompleted_counts):
                    height = bar.get_height()
                    center_y = bar.get_y() + height / 2
                    ax.annotate(f'{cc}/{uc}', xy=(bar.get_x() + bar.get_width() / 2, center_y),
                                xytext=(0, 0), textcoords="offset points",
                                ha='center', va='center')

        ax.set_xticks(x_positions)
        ax.set_xticklabels(list(car_number_to_index.keys())[start_index:end_index], rotation=90)

        handles, labels = ax.get_legend_handles_labels()
        unique_labels = set()
        good_handles = []
        good_labels = []

        for handle, label in zip(handles, labels):
            department = label.split(' ')[0]
            if department not in unique_labels:
                good_handles.append(handle)
                good_labels.append(department)
                unique_labels.add(department)

        ax.legend(good_handles, good_labels, loc='upper center', bbox_to_anchor=(0.5, 1.03),
                  fancybox=True, shadow=True, ncol=len(good_labels))
    plt.tight_layout(rect=[0, 0, 1, 1])
    plt.savefig(f"{book_name}{datetime.now().strftime('%Y%m%d%H%M%S')}_stacked_bar_charts.png", dpi=100)
    print("图片已保存")


if __name__ =='__main__':
    main()

