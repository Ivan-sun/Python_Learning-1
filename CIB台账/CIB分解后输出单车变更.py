import pandas as pd
from datetime import datetime
import numpy as np
import tkinter as tk
from tkinter import filedialog
import xlwings as xw
from openpyxl.utils import get_column_letter
from tqdm import tqdm

# 定义文件路径和工作表名称
file_path = filedialog.askopenfilename(title="选择Excel文件", filetypes=[("Excel Files", "*.xlsx;*.xls")])
sheet_name = 'Sheet1'

# 打开Excel文件并激活工作表
wb = xw.Book(file_path)
sht1 = wb.sheets[sheet_name]
book_name=sht1.range('A1').value
sht1.range('O6').value="制造分部\nMF" #修改原表中的制造分部df1中标题防止出现重复列


# 获取当前日期并格式化为年月日形式
current_date = datetime.now().strftime('%Y%m%d')
# 定义函数获取数据
def get_data_from_range(sheet, header_range, start_col_num, end_col_num):
    start_col = xw.utils.col_name(start_col_num)
    end_col = xw.utils.col_name(end_col_num)
    headers = sheet.range(header_range).value
    start_row = 10
    last_row = sheet.cells.last_cell.row
    end_row = sheet.range(f'{end_col}{last_row}').end('up').row
    data_range = sheet.range(f'{start_col}{start_row}:{end_col}{end_row}')
    data = data_range.value
    df = pd.DataFrame(data, columns=headers[:len(data[0])])
    df.dropna(how='all', inplace=True)
    return df

# 获取Dataframe1区域的数据
df1_1 = get_data_from_range(sht1, 'A6:S6', 1, 19)
df1_2 = get_data_from_range(sht1, 'W6:Y6', 23, 25)
df1_3 = get_data_from_range(sht1, 'T8:U8', 20, 21)
df1 = pd.concat([df1_1, df1_2,df1_3], axis=1)

# 提取合并单元格信息
def get_merge_info(cell):
    if cell.api.MergeCells:
        merge_area = cell.api.MergeArea
        return (
            merge_area.Cells(1).Row, 
            merge_area.Cells(1).Column, 
            merge_area.Cells(merge_area.Rows.Count, 1).Row, 
            merge_area.Cells(1, merge_area.Columns.Count).Column
        )
    else:
        return cell.row, cell.column, cell.row, cell.column

data = []
max_column = sht1.used_range.last_cell.column
for col_index in range(ord('Z') - ord('A'), max_column + 1):
    col_letter = get_column_letter(col_index + 1)
    cell = sht1[f'{col_letter}8']
    start_row_num, start_col_num, end_row_num, end_col_num = get_merge_info(cell)
    start_address = f"{get_column_letter(start_col_num)}{start_row_num}"
    end_address = f"{get_column_letter(end_col_num)}{end_row_num}"
    merge_count = (end_col_num - start_col_num + 1) * (end_row_num - start_row_num + 1) if cell.api.MergeCells else 1
    merge_context = cell.value if cell.api.MergeCells else ''
    data.append({
        'Start_Address': start_address,
        'End_Address': end_address,
        'Start_Column': start_col_num,
        'End_Column': end_col_num,
        'Merged_Cell_Count': merge_count,
        'Merge_Context': merge_context
    })

df_info = pd.DataFrame(data)
df_info.replace('', np.nan, inplace=True)
df_info.drop_duplicates(inplace=True)
df_info.dropna(subset=['Merge_Context'], inplace=True)

# 使用日期作为后缀创建新工作簿的名称
excel_file_name = f"{book_name}_SingleCar_{current_date}.xlsx"


# 初始化总的进度条
total_sheets = len(df_info)
with tqdm(total=total_sheets, desc="Writing Sheets", unit="sheets") as sheets_pbar:

    try:
        with pd.ExcelWriter(excel_file_name, engine='openpyxl') as writer:
            workbook = writer.book
            
            # 遍历每个工作表信息进行写入
            for index, row in df_info.iterrows():
                start_col_num = int(row['Start_Column'])
                end_col_num = int(row['End_Column'])
                merge_context = row['Merge_Context']
                df_i = get_data_from_range(sht1, 'Z9:AJ9', start_col_num, end_col_num)
                merged_df = pd.concat([df1.reset_index(drop=True), df_i.reset_index(drop=True)], axis=1)
                
                 # 定义需要移除的列名列表
                columns_to_drop = ['工艺\nMET', '外控\nSQC', '采购\nPROC', '运维\nO&M', '项目\nPM', '工程\nENG', 
                       '售后质量', 'Engineer', '质量-外控', '售后质量', '采购', '运维', '项目', '工程']
                # 移除指定的列
                merged_df.drop(columns=columns_to_drop, errors='ignore', inplace=True)

                # 计算不含标题的实际数据行数，并命名变量为 sum_CN（表示该项目所有的变更总数）
                sum_CN = merged_df.shape[0] - 1
                # print(merged_df.columns.tolist())


                # 筛选条件：删除'车间\nMF'、'外协\nOT'、'制造分部\nMF'列中全部都为“/”的行
                filter_condition = ~(merged_df[['车间\nMF', '外协\nOT', '制造分部\nMF']] == '/').all(axis=1)
                merged_df = merged_df.loc[filter_condition]

                # 计算不含标题的实际数据行数，并命名变量为 sum_MF_CN（表示该项目所有车间的变更总数）
                sum_MF_CN = merged_df.shape[0]
                # print(f'{sheet_name}的车间变更更总数： {sum_MF_CN}')

                # 筛选条件：筛选出'变更关闭状态' 列内容为‘open'的行
                merged_df = merged_df[merged_df['变更关闭状态'] == 'open']

                # 筛选条件更新：筛选出'生产'、'外协'、'制造分部'列中只要有一列的值包含”未完成“字符的行
                final_filter_condition = (merged_df[['生产', '外协\nOT', '制造分部\nMF']].apply(lambda x: x.str.contains('未完成')).any(axis=1))
                merged_df = merged_df.loc[final_filter_condition]
                # 筛选条件更新后，计算筛选完成的行数（排除标题行）
                sum_MF_OPEN_CN = merged_df.shape[0]




                # 写入数据到Excel
                merged_df.to_excel(writer, sheet_name=f"{merge_context}", index=False)
                
                # 更新工作表完成进度
                sheets_pbar.update(1)
                
    except Exception as e:
        print(f"写入文件时发生错误: {e}")

print(f"数据已成功写入到 {excel_file_name}")

# 关闭原工作簿，根据需要可取消注释
wb.save()
wb.app.quit()