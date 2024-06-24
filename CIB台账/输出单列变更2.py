import pandas as pd
from datetime import datetime
import numpy as np
import tkinter as tk
from tkinter import filedialog
import xlwings as xw
from openpyxl.utils import get_column_letter
from tqdm import tqdm

# 定义常量
FILE_TYPES = [("Excel Files", "*.xlsx;*.xls")]
ERROR_NO_FILE_SELECTED = "没有选择文件。"
ERROR_SHEET_NOT_VISIBLE = "错误：至少需要一个可见的工作表。请检查是否意外隐藏了工作表。"

def select_excel_file():
    """选择Excel文件的对话框"""
    return filedialog.askopenfilename(title="选择Excel文件", filetypes=FILE_TYPES)

def get_merge_info(cell):
    """获取合并单元格的信息"""
    if cell.api.MergeCells:
        merge_area = cell.api.MergeArea
        return (
            merge_area.Cells(1).Row,
            merge_area.Cells(1).Column,
            merge_area.Cells(merge_area.Rows.Count, 1).Row,
            merge_area.Cells(1, merge_area.Columns.Count).Column
        )
    else:
        return cell.row, cell.column, cell.row, cell.column

def get_data_from_range(sheet, header_range, start_col_num, end_col_num):
    start_col = xw.utils.col_name(start_col_num)
    end_col = xw.utils.col_name(end_col_num)
    headers = sheet.range(header_range).value
    start_row = 10
    last_row = sheet.cells.last_cell.row
    end_row = sheet.range(f'{end_col}{last_row}').end('up').row
    data_range = sheet.range(f'{start_col}{start_row}:{end_col}{end_row}')
    data = data_range.value
    df = pd.DataFrame(data, columns=headers[:len(data[0])])
    df.dropna(how='all', inplace=True)
    return df

def extract_merge_info(sheet, start_row, start_col):
    """提取合并单元格的信息，并过滤掉Merge_Context为空的记录"""
    max_column = sheet.used_range.last_cell.column
    data = []
    for col_index in range(ord(start_col) - ord('A'), max_column + 1):
        col_letter = get_column_letter(col_index + 1)
        cell = sheet[f'{col_letter}{start_row}']
        try:
            start_row_num, start_col_num, end_row_num, end_col_num = get_merge_info(cell)
            start_address = f"{get_column_letter(start_col_num)}{start_row_num}"
            end_address = f"{get_column_letter(end_col_num)}{end_row_num}"
            merge_count = (end_col_num - start_col_num + 1) * (end_row_num - start_row_num + 1) if cell.api.MergeCells else 1
            merge_context = cell.value if cell.api.MergeCells else ''
            if merge_context:
                data.append({
                    'Start_Address': start_address,
                    'End_Address': end_address,
                    'Start_Column': start_col_num,
                    'End_Column': end_col_num,
                    'Merged_Cell_Count': merge_count,
                    'Merge_Context': merge_context
                })
        except Exception as e:
            print(f"处理列 {col_letter} 时发生错误: {e}")
    return data

def main():
    try:
        file_path = select_excel_file()
        if not file_path:
            print(ERROR_NO_FILE_SELECTED)
            return

        sheet_name = 'Sheet1'
        wb = xw.Book(file_path)
        sht1 = wb.sheets[sheet_name]
        book_name = sht1.range('A1').value
        sht1.range('O6').value = "制造分部\nMF"

        df1_1 = get_data_from_range(sht1, 'A6:S6', 1, 19)
        df1_2 = get_data_from_range(sht1, 'W6:Y6', 23, 25)
        df1_3 = get_data_from_range(sht1, 'T8:U8', 20, 21)
        df1 = pd.concat([df1_1, df1_2, df1_3], axis=1)

        columns_to_drop = ['工艺\nMET', '外控\nSQC', '内控\nQC', '采购\nPROC', '运维\nO&M', '项目\nPM', '工程\nENG',
                           '售后质量', 'Engineer', '质量-外控', '质量-内控', '售后质量', '采购', '运维', '项目', '工程']
        df1.drop(columns=columns_to_drop, errors='ignore', inplace=True)

        current_date = datetime.now().strftime('%Y%m%d%h%%M')
        excel_file_name = f"{book_name}_TS_{current_date}.xlsx"

        data_TS = extract_merge_info(sht1, 6, 'Z')
        data_Car = extract_merge_info(sht1, 8, 'Z')

        df_TS_info = pd.DataFrame(data_TS)
        df_Car_info = pd.DataFrame(data_Car)

        # 遍历 df_Car_info,修改列名，增加车辆编号
        for index, row in df_Car_info.iterrows():
            start_address = f"{get_column_letter(row['Start_Column'])}9"
            end_address = f"{get_column_letter(row['End_Column'])}9"
            car_NO = row['Merge_Context']
            for i in sht1.range(f"{start_address}: {end_address}"):
                if car_NO not in i.value:
                    sht1.range(i).value = f"{car_NO}\n{sht1.range(i).value}"

        total_sheets = len(df_TS_info)
        with tqdm(total=total_sheets, desc="Writing Sheets:", unit="sheets") as sheets_pbar:
            try:
                with pd.ExcelWriter(excel_file_name, engine='openpyxl') as writer:
                    for index, row in df_TS_info.iterrows():
                        start_col_num = int(row['Start_Column'])
                        end_col_num = int(row['End_Column'])
                        merge_context = row['Merge_Context']
                        header_range = f"{get_column_letter(start_col_num)}9:{get_column_letter(end_col_num)}9"

                        df_i = get_data_from_range(sht1, header_range, start_col_num, end_col_num)
                        last_header = df_i.columns[-1]
                        if pd.isnull(last_header):
                            new_header = sht1.range(f"{get_column_letter(end_col_num)}7").value
                            df_i.rename(columns={last_header: new_header}, inplace=True)

                        merged_df = pd.concat([df1.reset_index(drop=True), df_i.reset_index(drop=True)], axis=1)
                        columns_to_remove = [col for col in merged_df.columns if any(substring in col for substring in columns_to_drop)]
                        merged_df.drop(columns=columns_to_remove, inplace=True, errors='ignore')

                        merged_df.dropna(subset=["列汇总"], inplace=True)
                        sum_CN = merged_df.shape[0] - 1

                        df_OPEN = merged_df[merged_df['列汇总'].str.contains('未完成')]
                        df_OPEN.to_excel(writer, sheet_name=f"{merge_context}", index=False)
                        sheets_pbar.update(1)
            except Exception as e:
                if "at least one sheet must be visible" in str(e):
                    print(ERROR_SHEET_NOT_VISIBLE)
                else:
                    raise e

    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    main()