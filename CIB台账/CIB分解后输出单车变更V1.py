import pandas as pd
from datetime import datetime
import numpy as np
import tkinter as tk
from tkinter import filedialog
import xlwings as xw
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font,Border,Side
from tqdm import tqdm

note ="""
    ┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄
                                    使用前注意事项
    0.本程序支持对Office Excel文件格式的变更处理,将变更汇总表按单车进行拆分
    1.被处理的CIB台帐请确保是从QMIS系统中自动导出的台账格式！

    2.本程序支持对多张台账进行拆分，拆分后的台账会自动保存到当前目录下
    
    3.如果拆分后发现在同一天对同一个项目的台账进行两次及以上，
    　使用时会出现报错，需要将第一次运行结果文件删除或者重新命名

    ┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄┄  
    """
print (note)
ss = input("====确认后，按任意键加回车继续====\n")



# 定义文件路径和工作表名称
file_path = filedialog.askopenfilename(title="选择Excel文件", filetypes=[("Excel Files", "*.xlsx;*.xls")])
sheet_name = 'Sheet1'

# 打开Excel文件并激活工作表
wb = xw.Book(file_path)
sht1 = wb.sheets[sheet_name]
book_name=sht1.range('A1').value
sht1.range('O6').value="制造分部\nMF" #修改原表中的制造分部df1中标题防止出现重复列


# 获取当前日期并格式化为年月日形式
current_date = datetime.now().strftime('%Y%m%d')
# 定义函数获取数据
def get_data_from_range(sheet, header_range, start_col_num, end_col_num):
    start_col = xw.utils.col_name(start_col_num)
    end_col = xw.utils.col_name(end_col_num)
    headers = sheet.range(header_range).value
    start_row = 10
    last_row = sheet.cells.last_cell.row
    end_row = sheet.range(f'{end_col}{last_row}').end('up').row
    data_range = sheet.range(f'{start_col}{start_row}:{end_col}{end_row}')
    data = data_range.value
    df = pd.DataFrame(data, columns=headers[:len(data[0])])
    df.dropna(how='all', inplace=True)
    return df

# 获取Dataframe1区域的数据
df1_1 = get_data_from_range(sht1, 'A6:S6', 1, 19)
df1_2 = get_data_from_range(sht1, 'W6:Y6', 23, 25)
df1_3 = get_data_from_range(sht1, 'T8:U8', 20, 21)
df1 = pd.concat([df1_1, df1_2,df1_3], axis=1)

# 提取合并单元格信息
def get_merge_info(cell):
    if cell.api.MergeCells:
        merge_area = cell.api.MergeArea
        return (
            merge_area.Cells(1).Row, 
            merge_area.Cells(1).Column, 
            merge_area.Cells(merge_area.Rows.Count, 1).Row, 
            merge_area.Cells(1, merge_area.Columns.Count).Column
        )
    else:
        return cell.row, cell.column, cell.row, cell.column

# 定义一个格式化的函数，便于复用
def apply_format(cell, font_name='等线', font_size=14, font_color='FF0000', is_bold=True, border=True, align_center=True):
    """应用格式到单元格"""
    font = Font(name=font_name, color=font_color, bold=is_bold, size=font_size)
    cell.font = font
    if border:
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        cell.border = thin_border
    if align_center:
        alignment = Alignment(horizontal='center', vertical='center')
        cell.alignment = alignment

data = []
max_column = sht1.used_range.last_cell.column
for col_index in range(ord('Z') - ord('A'), max_column + 1):
    col_letter = get_column_letter(col_index + 1)
    cell = sht1[f'{col_letter}8']
    start_row_num, start_col_num, end_row_num, end_col_num = get_merge_info(cell)
    start_address = f"{get_column_letter(start_col_num)}{start_row_num}"
    end_address = f"{get_column_letter(end_col_num)}{end_row_num}"
    merge_count = (end_col_num - start_col_num + 1) * (end_row_num - start_row_num + 1) if cell.api.MergeCells else 1
    merge_context = cell.value if cell.api.MergeCells else ''
    data.append({
        'Start_Address': start_address,
        'End_Address': end_address,
        'Start_Column': start_col_num,
        'End_Column': end_col_num,
        'Merged_Cell_Count': merge_count,
        'Merge_Context': merge_context
    })

df_info = pd.DataFrame(data)
df_info.replace('', np.nan, inplace=True)
df_info.drop_duplicates(inplace=True)
df_info.dropna(subset=['Merge_Context'], inplace=True)

# 使用日期作为后缀创建新工作簿的名称
excel_file_name = f"{book_name}_SingleCar_{current_date}.xlsx"


# 初始化总的进度条
total_sheets = len(df_info)
with tqdm(total=total_sheets, desc="Writing Sheets", unit="sheets") as sheets_pbar:

    try:
        with pd.ExcelWriter(excel_file_name, engine='openpyxl') as writer:
            workbook = writer.book
            
            # 遍历每个工作表信息进行写入
            for index, row in df_info.iterrows():
                start_col_num = int(row['Start_Column'])
                end_col_num = int(row['End_Column'])
                merge_context = row['Merge_Context']
                df_i = get_data_from_range(sht1, 'Z9:AJ9', start_col_num, end_col_num)
                merged_df = pd.concat([df1.reset_index(drop=True), df_i.reset_index(drop=True)], axis=1)
                
                 # 定义需要移除的列名列表
                columns_to_drop = ['工艺\nMET', '外控\nSQC', '采购\nPROC', '运维\nO&M', '项目\nPM', '工程\nENG', 
                       '售后质量', 'Engineer', '质量-外控', '售后质量', '采购', '运维', '项目', '工程']
                # 移除指定的列
                merged_df.drop(columns=columns_to_drop, errors='ignore', inplace=True)

                # 计算不含标题的实际数据行数，并命名变量为 sum_CN（表示该项目所有的变更总数）
                sum_CN = merged_df.shape[0] - 1
                

                # 筛选条件：删除'车间\nMF'、'外协\nOT'、'制造分部\nMF'列中全部都为“/”的行
                filter_condition = ~(merged_df[['车间\nMF', '外协\nOT', '制造分部\nMF']] == '/').all(axis=1)
                merged_df = merged_df.loc[filter_condition]

                # 计算不含标题的实际数据行数，并命名变量为 sum_MF_CN（表示该项目所有车间的变更总数）
                sum_MF_CN = merged_df.shape[0]
                

                # 筛选条件：筛选出'变更关闭状态' 列内容为‘open'的行
                merged_df = merged_df[merged_df['变更关闭状态'] == 'open']

                # 筛选条件更新：筛选出'生产'、'外协'、'制造分部'列中只要有一列的值包含”未完成“字符的行
                final_filter_condition = (merged_df[['生产', '外协\nOT', '制造分部\nMF']].apply(lambda x: x.str.contains('未完成')).any(axis=1))
                merged_df = merged_df.loc[final_filter_condition]
                # 筛选条件更新后，计算筛选完成的行数（排除标题行）
                sum_MF_OPEN_CN = merged_df.shape[0]

                

                # 写入数据到Excel
                merged_df.to_excel(writer, sheet_name=f"{merge_context}", startrow=2, startcol=0, index=False)
                
                # 获取当前活动的工作表
                activesheet = writer.sheets[f"{merge_context}"]
                # 合并单元格
                activesheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=activesheet.max_column)
                # 设置合并单元格的值
                activesheet.cell(row=1, column=1).value = f"{merge_context}车辆变更状态"
                activesheet.cell(row=2, column=1).value = "变更总数："
                activesheet.cell(row=2, column=2).value = sum_CN
                activesheet.cell(row=2, column=3).value = "车间变更数："
                activesheet.cell(row=2, column=4).value = sum_MF_CN
                activesheet.cell(row=2, column=5).value = "未关闭："
                activesheet.cell(row=2,column=6).value = sum_MF_OPEN_CN
                
                # 只调整第1到6列的列宽为16.5
                for col_index in range(1, 7):  # 从1到6，因为列索引是从1开始的
                    activesheet.column_dimensions[get_column_letter(col_index)].width = 16.5
                

                

                # 创建自定义的填充样式
                fill = PatternFill(start_color="ADD8E6",  # 浅蓝色的RGB代码
                                   end_color="ADD8E6",
                                   fill_type="solid")
                
                # 设置字体为黑体
                font = Font(name='黑体', bold=True, size=14)  # 注意：'SimHei'一般用于中文黑体，英文环境下可能需要其他字体名称
                
                # 设置单元格内容居中
                alignment = Alignment(horizontal="center", vertical="center")

                # 应用样式
                activesheet.cell(row=1, column=1).fill = fill
                activesheet.cell(row=1, column=1).font = font
                activesheet.cell(row=1, column=1).alignment = alignment

                # 应用格式到指定的单元格
                apply_format(activesheet.cell(row=1, column=1), font_size=16)  # 标题行可能需要更大字体
                apply_format(activesheet.cell(row=2, column=1))
                apply_format(activesheet.cell(row=2, column=2))
                apply_format(activesheet.cell(row=2, column=3))
                apply_format(activesheet.cell(row=2, column=4))
                apply_format(activesheet.cell(row=2, column=5))
                apply_format(activesheet.cell(row=2, column=6))
                
                # 更新工作表完成进度
                sheets_pbar.update(1)
                
    except Exception as e:
        print(f"写入文件时发生错误: {e}")

print(f"数据已成功写入到 {excel_file_name}")

# 关闭原工作簿，根据需要可取消注释
wb.save()
wb.app.quit()