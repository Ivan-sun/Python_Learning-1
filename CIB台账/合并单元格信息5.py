import pandas as pd
import tkinter as tk
from tkinter import filedialog
import xlwings as xw
from openpyxl.utils import get_column_letter

# 定义文件路径和工作表名称
# file_path = r'CIB台账\巴西圣保罗项目单轨项目.xlsx'
file_path = filedialog.askopenfilename(title="选择Excel文件", filetypes=[("Excel Files", "*.xlsx;*.xls")])
sheet_name = 'Sheet1'

# 打开Excel文件并激活工作表
wb = xw.Book(file_path)
sht1 = wb.sheets[sheet_name]

# 定义起始行和列
start_row = 8
start_col = 'Z'

# 函数来获取合并单元格的信息
def get_merge_info(cell):
    if cell.api.MergeCells:
        merge_area = cell.api.MergeArea
        return (
            merge_area.Cells(1).Row,  # Start_Row
            merge_area.Cells(1).Column,  # Start_Column
            merge_area.Cells(merge_area.Rows.Count, 1).Row,  # End_Row
            merge_area.Cells(1, merge_area.Columns.Count).Column  # End_Column
        )
    else:
        return cell.row, cell.column, cell.row, cell.column

# 获取工作表的最大列号
max_column = sht1.used_range.last_cell.column

# 提取指定行的数据，并增加起始地址和结束地址列
data = []
for col_index in range(ord(start_col) - ord('A'), max_column + 1):
    col_letter = get_column_letter(col_index + 1)
    cell = sht1[f'{col_letter}{start_row}']
    
    # 获取合并单元格的起始和结束行列信息
    start_row_num, start_col_num, end_row_num, end_col_num = get_merge_info(cell)
    
    # 计算起始和结束地址
    start_address = f"{get_column_letter(start_col_num)}{start_row_num}"
    end_address = f"{get_column_letter(end_col_num)}{end_row_num}"
    
    # 获取并记录合并单元格数量，以及合并内容
    merge_count = (end_col_num - start_col_num + 1) * (end_row_num - start_row_num + 1) if cell.api.MergeCells else 1
    merge_context = cell.value if cell.api.MergeCells else ''
    
    data.append({
        'Start_Address': start_address,
        'End_Address': end_address,
        'Start_Column': start_col_num,
        'End_Column': end_col_num,
        'Merged_Cell_Count': merge_count,
        'Merge_Context': merge_context
    })

# 将信息转换为DataFrame
df_info = pd.DataFrame(data)

# 去除重复行并筛选非空的'Merge_Context'
df_info = df_info.drop_duplicates().dropna(subset=['Merge_Context'])
print(df_info)
'''  如果需要存储则需要从此处取消注释
# 检查并删除已存在的'Merge_Info'工作表
print("尝试删除 'Merge_Info'...")
if 'Merge_Info' in wb.sheet_names:
    wb.sheets['Merge_Info'].delete()
    print("'Merge_Info' 工作表已删除。")
else:
    print("'Merge_Info' 工作表未找到。")

# 添加新的'Merge_Info'工作表
sht2 = wb.sheets.add(after=sht1, name='Merge_Info')

# 将筛选后的DataFrame写入新工作表
sht2.range('A1').options(index=False).value = df_info
'''
# 保存并关闭工作簿
wb.save()
wb.app.quit()





for index, row in df_info.iterrows():
    start_col_num = row['Start_Column']
    end_col_num = row['End_Column']
    merge_context = row['Merge_Context']
    
    print(f"开始列号：{start_col_num} ,结束列号；{end_col_num} ,车辆编号{merge_context}")
    

# -------



# -------------------
